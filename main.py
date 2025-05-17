import sys
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QLabel,
    QVBoxLayout,
    QFrame,
    QPushButton,
    QFileDialog,
)
from PyQt5.QtCore import Qt, QTimer
from snipping import Snip
import keyboard  # For global key tracking
import threading  # For background thread to handle global key events
import time  # To use time.sleep()
import webbrowser  # Import the webbrowser module

from openpyxl import Workbook, load_workbook  # Add this import

class ScreenRegionSelector(QMainWindow):

    def __init__(self):
        super().__init__(None)
        self.m_width = 400
        self.m_height = 500

        self.setWindowTitle("Kantan-ji: The Kanji Simplifier")
        self.setMinimumSize(self.m_width, self.m_height)

        frame = QFrame()
        frame.setContentsMargins(0, 0, 0, 0)
        lay = QVBoxLayout(frame)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.setContentsMargins(5, 5, 5, 5)

        self.label = QLabel()
        self.btn_capture = QPushButton("Capture")
        self.btn_capture.clicked.connect(self.capture)

        self.btn_copy = QPushButton("Copy to Clipboard")
        self.btn_copy.clicked.connect(self.copy)
        self.btn_copy.setVisible(False)

        self.btn_open_browser = QPushButton("Open in Browser")
        self.btn_open_browser.clicked.connect(self.open_in_browser)
        self.btn_open_browser.setVisible(False)

        self.btn_add_flashcard = QPushButton("Add to Flashcard Spreadsheet")
        self.btn_add_flashcard.clicked.connect(self.add_to_flashcard)
        self.btn_add_flashcard.setVisible(False)

        lay.addWidget(self.label)
        lay.addWidget(self.btn_capture)
        lay.addWidget(self.btn_copy)
        lay.addWidget(self.btn_open_browser)
        lay.addWidget(self.btn_add_flashcard)

        self.setCentralWidget(frame)

        # Set to track key presses
        self.pressed_keys = set()

        # Start global key listener
        keyboard.on_press(self.global_key_press)
        keyboard.on_release(self.global_key_release)

        # Flag to track whether the capture window is shown
        self.capturing = False

        # Start background thread to listen for Esc key globally
        self.app_is_running = True
        self.key_thread = threading.Thread(target=self.key_handler, daemon=True)
        self.key_thread.start()

        # Store the extracted text here
        self.extracted_text = ""

    def global_key_press(self, event):
        if event.name == '[':
            self.pressed_keys.add('[')
        elif event.name == ']':
            self.pressed_keys.add(']')

        if '[' in self.pressed_keys and ']' in self.pressed_keys:
            QTimer.singleShot(0, self.capture)
            self.pressed_keys.clear()

    def global_key_release(self, event):
        if event.name == '[':
            self.pressed_keys.discard('[')
        elif event.name == ']':
            self.pressed_keys.discard(']')

    def capture(self):
        self.showMinimized()
        self.capturer = Snip(self)
        self.capturer.show()
        self.capturing = True

        self.btn_copy.setVisible(True)
        self.btn_open_browser.setVisible(True)
        self.btn_add_flashcard.setVisible(True)

    def copy(self):
        if self.extracted_text:
            clipboard = QApplication.clipboard()
            clipboard.setText(self.extracted_text)
            print(f"Text copied to clipboard: {self.extracted_text}")
        else:
            print("No text to copy!")

    def open_in_browser(self):
        if self.extracted_text:
            extracted_text = self.extracted_text.split("\n")[1]
            url = f"https://jisho.org/search/{extracted_text}"
            webbrowser.open(url)
            print(f"Text opened in browser: {extracted_text}")
        else:
            print("No text to open in browser!")

    def add_to_flashcard(self):
        """Add the extracted text and simplified text to a spreadsheet."""
        if not self.extracted_text:
            print("No text to add to flashcard!")
            return

        try:
            lines = self.extracted_text.split("\n")
            original = lines[1] if len(lines) > 1 else ""
            simplified = lines[3] if len(lines) > 3 else ""
        except Exception as e:
            print(f"Error parsing text: {e}")
            return

        filename = "flashcards.xlsx"
        try:
            try:
                wb = load_workbook(filename)
                ws = wb.active
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
                ws.append(["Original", "Simplified"])

            ws.append([original, simplified])
            wb.save(filename)
            print(f"Added to flashcard spreadsheet: {original} | {simplified}")
        except Exception as e:
            print(f"Error writing to spreadsheet: {e}")

    def close_capture_window(self):
        if self.capturer:
            self.capturer.hide()
        self.capturing = False
        self.show()

    def key_handler(self):
        while self.app_is_running:
            if keyboard.is_pressed("esc"):
                self.close_capture_window()
            time.sleep(0.1)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet("""
    QFrame {
        background-color: #3f3f3f;
    }
    QPushButton {
        border-radius: 5px;
        background-color: rgb(60, 90, 255);
        padding: 10px;
        color: white;
        font-weight: bold;
        font-family: Arial;
        font-size: 12px;
    }
    QPushButton::hover {
        background-color: rgb(60, 20, 255)
    }
    """)
    selector = ScreenRegionSelector()
    selector.show()
    app.exit(app.exec_())
